import streamlit as st
import pandas as pd
import io, unicodedata
from datetime import datetime
from database import (listar_usuarios, criar_usuario, editar_usuario,
                      alterar_senha, listar_unidades, criar_unidade,
                      get_todas_metas, set_meta, resetar_projetos_teste,
                      listar_projetos, deletar_usuario,
                      importar_projetos_lote, normalizar_valor_lista, lancar_real,
                      exportar_backup_completo, restaurar_backup_completo,
                      TIPOS_PROJETO, VA_GGF_OPTS, STATUS_OPTS, PERFIS_LBL)

def render(user, **colors):
    NAVY=colors.get("NAVY","#0B0F2B"); RED=colors.get("RED","#D93B3B")
    if user["perfil"] != "admin":
        st.error("⛔ Acesso restrito a administradores."); return

    st.markdown('<span class="st">Administração</span>', unsafe_allow_html=True)
    tab_users, tab_editar, tab_metas, tab_unid, tab_senha, tab_backup, tab_import, tab_reset = st.tabs([
        "👥 Usuários","✏️ Editar Usuário","🎯 Metas","🏭 Unidades","🔑 Senhas",
        "💾 Backup","📥 Importar Excel","🗑️ Reset"])

    with tab_users:
        usuarios = listar_usuarios()
        rows = "".join(f"""<tr>
          <td style="font-size:11px;font-weight:600;">{u['nome']}</td>
          <td style="font-size:11px;">{u['email']}</td>
          <td style="font-size:11px;">{PERFIS_LBL.get(u['perfil'],u['perfil'])}</td>
          <td style="font-size:11px;">{u.get('unidade') or '— Global'}</td>
          <td style="font-size:11px;">{'✅' if u['ativo'] else '❌'}</td>
        </tr>""" for u in usuarios)
        st.markdown(f"""<table class="dt"><thead><tr>
          <th>Nome</th><th>E-mail</th><th>Perfil</th><th>Unidade</th><th>Ativo</th>
        </tr></thead><tbody>{rows}</tbody></table>""", unsafe_allow_html=True)

        st.markdown("---")
        st.markdown("**Criar novo usuário**")
        unidades = listar_unidades()
        nomes_u  = ["— Acesso Global"] + [u["nome"] for u in unidades]
        with st.form("form_user", clear_on_submit=True):
            c1,c2=st.columns(2)
            with c1:
                nome_u  = st.text_input("Nome completo *")
                email_u = st.text_input("E-mail *")
                senha_u = st.text_input("Senha inicial *", type="password")
            with c2:
                perfil_u  = st.selectbox("Perfil *", list(PERFIS_LBL.keys()),
                                          format_func=lambda x: PERFIS_LBL[x])
                unidade_u = st.selectbox(
                    "Unidade (Global para admin/cost_control/visualizador global)",
                    nomes_u)
            if st.form_submit_button("➕ Criar Usuário", use_container_width=True):
                if not nome_u or not email_u or not senha_u:
                    st.error("Preencha todos os campos.")
                else:
                    unid = None if unidade_u=="— Acesso Global" else unidade_u
                    try:
                        criar_usuario(nome_u,email_u,senha_u,perfil_u,unid)
                        st.success(f"✅ **{nome_u}** criado!"); st.rerun()
                    except Exception as e: st.error(f"Erro: {e}")

    with tab_editar:
        usuarios = listar_usuarios()
        unidades = listar_unidades()
        nomes_u2 = ["— Acesso Global"] + [u["nome"] for u in unidades]
        sel_u = st.selectbox("Selecionar:",[u["nome"] for u in usuarios],key="ed_user")
        u_sel = next(u for u in usuarios if u["nome"]==sel_u)
        with st.form("form_ed_user"):
            c1,c2=st.columns(2)
            with c1:
                novo_nome  = st.text_input("Nome", value=u_sel["nome"])
                novo_email = st.text_input("E-mail", value=u_sel["email"])
                ativo      = st.checkbox("Ativo", value=bool(u_sel["ativo"]))
            with c2:
                novo_perfil = st.selectbox("Perfil", list(PERFIS_LBL.keys()),
                    index=list(PERFIS_LBL.keys()).index(u_sel["perfil"])
                    if u_sel["perfil"] in PERFIS_LBL else 0,
                    format_func=lambda x: PERFIS_LBL[x])
                unid_atual = u_sel.get("unidade") or "— Acesso Global"
                idx_unid   = nomes_u2.index(unid_atual) if unid_atual in nomes_u2 else 0
                nova_unid  = st.selectbox("Unidade", nomes_u2, index=idx_unid)
            if st.form_submit_button("💾 Salvar", use_container_width=True):
                editar_usuario(u_sel["id"],{
                    "nome":novo_nome,"email":novo_email.lower(),
                    "perfil":novo_perfil,
                    "unidade":None if nova_unid=="— Acesso Global" else nova_unid,
                    "ativo":int(ativo)})
                st.success(f"✅ **{novo_nome}** atualizado!"); st.rerun()

        st.markdown("---")
        st.markdown("**🗑️ Excluir usuário**")
        admins_ativos = [u for u in usuarios if u["perfil"]=="admin" and u["ativo"]]
        if u_sel["id"] == user["id"]:
            st.info("Você não pode excluir seu próprio usuário logado.")
        elif u_sel["perfil"]=="admin" and u_sel["ativo"] and len(admins_ativos) <= 1:
            st.info("Esse é o único administrador ativo — não pode ser excluído.")
        else:
            conf_del = st.checkbox(
                f"Confirmo que quero excluir **{u_sel['nome']}** permanentemente.",
                key=f"confdel_{u_sel['id']}")
            if st.button("🗑️ Excluir Usuário", disabled=not conf_del,
                        key=f"btndel_{u_sel['id']}"):
                if deletar_usuario(u_sel["id"]):
                    st.success(f"✅ **{u_sel['nome']}** excluído."); st.rerun()
                else:
                    st.error(
                        f"Não foi possível excluir **{u_sel['nome']}**: existem projetos "
                        f"ou lançamentos no histórico vinculados a ele. Desative-o "
                        f"(campo Ativo acima) em vez de excluir.")

    with tab_metas:
        st.markdown("**Definir Meta Anual por Unidade / Área**")
        anos = list(range(2026,2031))
        ano_sel = st.selectbox("Ano:",anos,key="adm_ano")
        metas = get_todas_metas(ano_sel)
        with st.form("form_metas"):
            vals={}
            for m in metas:
                vals[m["nome"]] = st.number_input(
                    f"{m['nome']} ({m['tipo']})",
                    value=float(m["valor"]),step=10000.0,format="%.2f",
                    key=f"meta_{m['nome']}_{ano_sel}")
            if st.form_submit_button("💾 Salvar Metas", use_container_width=True):
                for nome,val in vals.items(): set_meta(nome,ano_sel,val)
                st.success("✅ Metas salvas!")

    with tab_unid:
        unidades = listar_unidades(so_ativas=False)
        rows_u="".join(f"""<tr>
          <td style="font-size:11px;font-weight:600;">{u['nome']}</td>
          <td style="font-size:11px;">{u['tipo']}</td>
          <td style="font-size:11px;">{'✅ Ativa' if u['ativo'] else '❌ Inativa'}</td>
        </tr>""" for u in unidades)
        st.markdown(f"""<table class="dt">
          <thead><tr><th>Nome</th><th>Tipo</th><th>Status</th></tr></thead>
          <tbody>{rows_u}</tbody></table>""", unsafe_allow_html=True)
        st.markdown("---")
        with st.form("form_unid", clear_on_submit=True):
            c1,c2=st.columns(2)
            with c1: nome_nu=st.text_input("Nome *")
            with c2: tipo_nu=st.selectbox("Tipo",["planta","area"])
            if st.form_submit_button("➕ Criar", use_container_width=True):
                if nome_nu: criar_unidade(nome_nu,tipo_nu); st.success(f"✅ **{nome_nu}** criada!"); st.rerun()

    with tab_senha:
        usuarios=listar_usuarios()
        sel_s=st.selectbox("Usuário:",[u["nome"] for u in usuarios],key="adm_senha")
        u_s=next(u for u in usuarios if u["nome"]==sel_s)
        with st.form("form_senha"):
            nova=st.text_input("Nova senha",type="password")
            conf=st.text_input("Confirmar",type="password")
            if st.form_submit_button("🔑 Alterar Senha"):
                if nova!=conf: st.error("Senhas não conferem.")
                elif len(nova)<6: st.error("Mínimo 6 caracteres.")
                else: alterar_senha(u_s["id"],nova); st.success(f"✅ Senha alterada!")

    with tab_backup:
        st.markdown("**💾 Backup Completo da Plataforma**")
        st.caption(
            "Esse é o fluxo recomendado pra manter tudo seguro enquanto o banco ainda não é "
            "persistente: baixe o backup periodicamente. Se a base zerar, suba o último "
            "backup baixado e a plataforma volta pro ponto exato de onde parou — projetos, "
            "checklist, validação, ganho único, real lançado mês a mês, metas e usuários.")

        st.markdown("##### ⬇️ Baixar backup")
        linhas_proj, linhas_metas, linhas_usuarios = exportar_backup_completo()
        n_meses_real = len([c for c in (linhas_proj[0].keys() if linhas_proj else []) if c.startswith("Real ")])
        st.caption(f"Agora mesmo: {len(linhas_proj)} projeto(s) (com até {n_meses_real} mês(es) de "
                  f"real cada), {len(linhas_metas)} meta(s), {len(linhas_usuarios)} usuário(s) cadastrado(s).")

        def _gerar_backup_xlsx():
            buf = io.BytesIO()
            with pd.ExcelWriter(buf, engine="openpyxl") as writer:
                pd.DataFrame(linhas_proj).to_excel(writer, sheet_name="Projetos", index=False)
                pd.DataFrame(linhas_metas).to_excel(writer, sheet_name="Metas", index=False)
                pd.DataFrame(linhas_usuarios).to_excel(writer, sheet_name="Usuarios", index=False)
            buf.seek(0)
            return buf

        st.download_button("⬇️ Baixar Backup Completo (.xlsx)", data=_gerar_backup_xlsx(),
            file_name=f"backup_plataforma_delga_{datetime.now().strftime('%Y-%m-%d')}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True, type="primary")
        st.caption("A aba 'Usuarios' guarda a senha num formato técnico (hash) — não dá pra ler a "
                  "senha ali, só serve pra restaurar o login sem precisar redefinir depois. Não edite essa coluna.")

        st.markdown("---")
        st.markdown("##### ⬆️ Restaurar backup")
        st.warning("⚠️ Restaurar **apaga todos os projetos e metas atuais** e recarrega exatamente "
                  "o que estiver no arquivo. Usuários que já existirem (mesmo e-mail) não são "
                  "apagados nem duplicados — só os que faltarem são recriados. Não soma com o que já existe.")
        arq_backup = st.file_uploader("Enviar arquivo de backup (.xlsx)", type=["xlsx"], key="up_backup")

        if arq_backup is not None:
            try:
                xls_b = pd.read_excel(arq_backup, sheet_name=None)
                df_p = xls_b.get("Projetos")
                df_m = xls_b.get("Metas")
                df_u = xls_b.get("Usuarios")
                if df_p is None:
                    st.error("Esse arquivo não parece um backup da plataforma — falta a aba 'Projetos'.")
                else:
                    linhas_proj_r = df_p.fillna("").to_dict("records")
                    linhas_metas_r = df_m.fillna("").to_dict("records") if df_m is not None else []
                    linhas_usuarios_r = df_u.fillna("").to_dict("records") if df_u is not None else []
                    n_meses_r = len([c for c in linhas_proj_r[0].keys() if str(c).startswith("Real ")]) if linhas_proj_r else 0
                    st.markdown(f"**Prévia:** {len(linhas_proj_r)} projeto(s) (com colunas de real pra "
                              f"até {n_meses_r} mês(es)), {len(linhas_metas_r)} meta(s), "
                              f"{len(linhas_usuarios_r)} usuário(s) nesse arquivo — isso vai **substituir** "
                              f"os projetos e metas atuais.")
                    confirmar_rest = st.checkbox(
                        "Confirmo que quero apagar os dados atuais e restaurar este backup.")
                    if st.button("♻️ Restaurar Backup", disabled=not confirmar_rest,
                                type="primary", use_container_width=True):
                        n_p, n_l, n_m, n_u, erros_r = restaurar_backup_completo(
                            linhas_proj_r, linhas_metas_r, linhas_usuarios_r, user["id"])
                        st.success(f"✅ Restaurado: {n_p} projeto(s), {n_l} lançamento(s) de real, "
                                  f"{n_m} meta(s) e {n_u} usuário(s) novo(s) recriado(s).")
                        if erros_r:
                            st.error(f"{len(erros_r)} projeto(s) falharam: " +
                                    "; ".join(f"{e['nome']} ({e['erro']})" for e in erros_r))
                        st.rerun()
            except Exception as e:
                st.error(f"Não consegui ler esse arquivo: {e}")

    with tab_import:
        st.markdown("**📥 Importação em Lote — Projetos e Metas (nível de arranque)**")
        st.caption(
            "Use isso pra subir de uma vez os projetos que já existem fora da plataforma "
            "(planilha de controle). O lançamento de real mês a mês continua manual, "
            "projeto por projeto, em Controle de Custos.")

        unidades_disp = [u["nome"] for u in listar_unidades()]

        # ── Gerar modelo ────────────────────────────────────────────────────
        def _gerar_modelo():
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill
            from openpyxl.worksheet.datavalidation import DataValidation
            wb = Workbook()
            ws1 = wb.active; ws1.title = "Projetos"
            cols1 = ["Unidade","Tipo","VA/GGF","Nome do Projeto","Descrição","Responsável",
                     "Data Início","Data Fim","Valor Previsto","Mês Primeiro Retorno",
                     "Validação","Valor Calculado Custos","Status"]
            ws1.append(cols1)
            for c in range(1,len(cols1)+1):
                cell = ws1.cell(row=1,column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0B0F2B")
            exemplo = ["Diadema","BSW","VA","Projeto Exemplo — apagar esta linha",
                       "Descrição do projeto","Fulano de Tal","2026-01-15","2026-06-30",
                       120000,"2026-09-01","OK",40000,"Em Execução"]
            ws1.append(exemplo)
            for c in range(1,len(cols1)+1):
                ws1.cell(row=2,column=c).font = Font(italic=True, color="8A9BB0")
            widths = [14,14,10,32,30,18,13,13,14,18,11,20,14]
            for i,w in enumerate(widths,start=1):
                ws1.column_dimensions[chr(64+i) if i<=26 else "A"].width = w

            def _add_dv(col_letter, opcoes, n_rows=200):
                dv = DataValidation(type="list", formula1=f'"{",".join(opcoes)}"', allow_blank=True)
                ws1.add_data_validation(dv)
                dv.add(f"{col_letter}2:{col_letter}{n_rows}")
            _add_dv("A", unidades_disp)
            _add_dv("B", TIPOS_PROJETO)
            _add_dv("C", VA_GGF_OPTS)
            _add_dv("J", ["OK","NOK","Pendente"])
            _add_dv("M", STATUS_OPTS)

            ws2 = wb.create_sheet("Metas")
            ws2.append(["Unidade","Ano","Valor da Meta"])
            for c in range(1,4):
                cell = ws2.cell(row=1,column=c)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor="0B0F2B")
            ws2.append(["Diadema",2026,300000])
            ws2.cell(row=2,column=1).font = Font(italic=True, color="8A9BB0")
            ws2.column_dimensions["A"].width=14; ws2.column_dimensions["C"].width=16
            dv2 = DataValidation(type="list", formula1=f'"{",".join(unidades_disp)}"', allow_blank=True)
            ws2.add_data_validation(dv2); dv2.add("A2:A200")

            buf = io.BytesIO(); wb.save(buf); buf.seek(0)
            return buf

        c_dl, _ = st.columns([1,3])
        with c_dl:
            st.download_button("⬇️ Baixar modelo (.xlsx)", data=_gerar_modelo(),
                file_name="modelo_importacao_delga.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True)
        st.caption("Apague a linha de exemplo antes de preencher. As colunas Unidade, Tipo, "
                   "VA/GGF, Validação e Status têm lista suspensa pra evitar erro de digitação. "
                   "Se você já tem a planilha oficial de Controle de Indicadores da Delga (uma "
                   "aba por unidade), pode subir ela direto — não precisa reformatar.")

        st.markdown("---")
        arquivo = st.file_uploader("Enviar planilha preenchida (.xlsx)", type=["xlsx"])

        def _norm(s):
            s = str(s or "").strip().lower()
            return "".join(c for c in unicodedata.normalize("NFD", s) if unicodedata.category(c) != "Mn")

        def _tokens(s):
            import re
            s = _norm(s).replace("º","").replace("ª","")
            s = re.sub(r"[^a-z0-9]+"," ",s)
            return set(t for t in s.split() if t)

        def _parse_data(v):
            if v is None or (hasattr(pd,"isna") and pd.isna(v)) or v == "": return ""
            if isinstance(v,(pd.Timestamp,datetime)): return v.strftime("%Y-%m-%d")
            try: return datetime.strptime(str(v)[:10],"%Y-%m-%d").strftime("%Y-%m-%d")
            except: return ""

        def _parse_mes(v):
            if v is None or (hasattr(pd,"isna") and pd.isna(v)) or v == "": return None
            if isinstance(v,(pd.Timestamp,datetime)): return v.strftime("%Y-%m-01")
            try: return datetime.strptime(str(v)[:7],"%Y-%m").strftime("%Y-%m-01")
            except: return None

        # ── Campos reconhecidos e seus sinônimos (usado nos dois modos) ───────
        CAMPOS = [
            ("unidade","Unidade",True,
             ["unidade","planta","area","fabrica","unid","filial","site"]),
            ("tipo","Tipo",True,
             ["tipo","categoria","classificacao","pilar","tipo projeto","tipo de projeto","natureza"]),
            ("nome","Nome do Projeto",True,
             ["nome do projeto","nome projeto","projeto","nome","titulo","iniciativa"]),
            ("va_ggf","VA/GGF",False,
             ["va ggf","va/ggf","vaggf","va","ggf","material auxiliar","mat aux","classificacao va"]),
            ("descricao","Descrição",False,
             ["descricao","objetivo","detalhamento","obs","observacao","escopo"]),
            ("responsavel","Responsável",False,
             ["responsavel","owner","dono","facilitador","gestor","encarregado"]),
            ("inicio","Data Início",False,
             ["data inicio","inicio","dt inicio","data de inicio","abertura"]),
            ("termino","Data Fim",False,
             ["data fim","fim","termino","data termino","dt fim","conclusao","data conclusao","encerramento"]),
            ("previsto_unidade","Valor Previsto",True,
             ["valor previsto","previsto","valor estimado","estimado"]),
            ("mes_primeiro_retorno","Mês Primeiro Retorno",False,
             ["mes primeiro retorno","mes do primeiro retorno","ganho a partir de","primeiro retorno",
              "1o retorno","1 retorno","mes retorno","data retorno","retorno previsto"]),
            ("validacao","Validação",False,
             ["validacao","validador","aprovacao","ok nok","validado custos","aprovado custos"]),
            ("valor_custos_bruto","Valor Calculado Custos",False,
             ["valor calculado custos","valor calculado por custos","calculado custos","saving custos",
              "valor custos","custos calculado","saving","valor validado custos","calc custos"]),
            ("status","Status",False,
             ["status","situacao","fase","andamento","etapa"]),
        ]

        def _melhor_coluna(colunas, aliases):
            """Acha, entre uma lista de nomes de coluna, a que melhor bate com
            os sinônimos de um campo — por pontuação de sobreposição de palavras."""
            melhor, melhor_score = None, 0
            for c in colunas:
                cn = _norm(str(c))
                cn_tok = _tokens(str(c))
                score = 0
                for a in aliases:
                    if a == cn: score = max(score, 100)
                    elif a in cn: score = max(score, 70)
                    a_tok = _tokens(a)
                    if a_tok and cn_tok:
                        overlap = len(a_tok & cn_tok) / len(a_tok | cn_tok)
                        score = max(score, int(overlap*60))
                if score > melhor_score:
                    melhor_score, melhor = score, str(c)
            return melhor if melhor_score >= 25 else None

        def _linha_de(row, mapa, linha_num, unidade_fixa=None):
            """Extrai e valida uma linha de projeto a partir de um mapeamento
            de colunas já resolvido. unidade_fixa força a unidade (modo Delga,
            onde a unidade vem do nome da aba, não de uma coluna)."""
            def _get(chave):
                col = mapa.get(chave)
                return row[col] if col else None
            nome = _get("nome")
            if nome is None or (hasattr(pd,"isna") and pd.isna(nome)) or not str(nome).strip():
                return None
            motivo = []
            unid = unidade_fixa or normalizar_valor_lista(_get("unidade"), unidades_disp)
            if not unid: motivo.append("unidade não reconhecida")
            tipo = normalizar_valor_lista(_get("tipo"), TIPOS_PROJETO)
            if not tipo: return None  # provavelmente linha de outra tabela, não de projeto
            va = normalizar_valor_lista(_get("va_ggf"), VA_GGF_OPTS) or "VA"
            mpr = _parse_mes(_get("mes_primeiro_retorno"))
            prev = _get("previsto_unidade")
            try:
                prev = float(prev) if prev is not None and not (hasattr(pd,"isna") and pd.isna(prev)) else 0.0
            except (ValueError, TypeError):
                prev = 0.0
            if prev != prev: prev = 0.0  # NaN nunca é igual a si mesmo
            if prev <= 0: motivo.append("valor previsto zerado")
            valid = normalizar_valor_lista(_get("validacao"), ["OK","NOK","Pendente"]) or "Pendente"
            status = normalizar_valor_lista(_get("status"), STATUS_OPTS)
            linha = {
                "linha_planilha": linha_num, "nome": str(nome).strip(),
                "unidade": unid, "tipo": tipo, "va_ggf": va,
                "descricao": str(_get("descricao") or ""),
                "responsavel": str(_get("responsavel") or ""),
                "inicio": _parse_data(_get("inicio")),
                "termino": _parse_data(_get("termino")),
                "previsto_unidade": prev,
                "mes_primeiro_retorno": mpr,
                "validacao": valid,
                "valor_custos_bruto": _get("valor_custos_bruto"),
                "status": status,
            }
            if motivo: linha["_erro"] = ", ".join(motivo)
            return linha

        if arquivo is not None:
            try:
                xls = pd.read_excel(arquivo, sheet_name=None, header=None)
            except Exception as e:
                st.error(f"Não consegui ler o arquivo: {e}"); xls = None

            linhas_ok, linhas_erro, metas_linhas = [], [], []

            if xls is not None:
                nomes_abas = list(xls.keys())

                # ── MODO 1: formato oficial Delga (uma aba por unidade) ────────
                UNIDADES_CONHECIDAS = ["Diadema","Jarinu","Ferraz","Anchieta","São Leopoldo",
                                       "Compras","Vendas","Corporativo"]
                mapa_abas_unidade = {}
                for aba in nomes_abas:
                    unid_match = normalizar_valor_lista(aba, UNIDADES_CONHECIDAS)
                    if unid_match:
                        mapa_abas_unidade[unid_match] = aba

                modo_delga = len(mapa_abas_unidade) >= 3
                if modo_delga:
                    st.success(f"📋 Reconheci o formato oficial de Controle de Indicadores da Delga — "
                              f"{len(mapa_abas_unidade)} unidade(s) encontrada(s): "
                              f"{', '.join(mapa_abas_unidade.keys())}")

                    mapas_cache = {}  # {n_colunas: {"campos_idx":{chave:idx}, "mes_idx":{idx:(ano,mes)}}}
                    pendentes_sem_header = []

                    def _processar_aba(unidade, aba, mapa_local, col_mes, df, primeira_linha_dados):
                        for i,(ridx,row) in enumerate(df.iterrows()):
                            linha = _linha_de(row, mapa_local, primeira_linha_dados+1+i, unidade_fixa=unidade)
                            if linha is None: continue
                            linha["_aba"] = aba
                            # Linha "Real" fica imediatamente abaixo da linha do
                            # projeto (planilha vem com 2 linhas por projeto).
                            # Acha a coluna marcadora "Previsto/Real" olhando a
                            # própria linha do projeto por essa palavra exata.
                            real_mensal = {}
                            row_atual = primeira_linha_dados+i
                            marcador_col = None
                            for c in range(df_raw.shape[1]):
                                if df_raw.iat[row_atual, c] == "Previsto":
                                    marcador_col = c; break
                            if marcador_col is not None and row_atual+1 < len(df_raw):
                                if df_raw.iat[row_atual+1, marcador_col] == "Real":
                                    for c,(ano_m,mes_m) in col_mes.items():
                                        v = df_raw.iat[row_atual+1, c]
                                        if v is not None and not (hasattr(pd,"isna") and pd.isna(v)):
                                            try: real_mensal[(ano_m,mes_m)] = float(v)
                                            except (ValueError,TypeError): pass
                            linha["_real_mensal"] = real_mensal
                            if "_erro" in linha: linhas_erro.append(linha)
                            else: linhas_ok.append(linha)

                    # Passo 1: abas com cabeçalho de texto normal
                    for unidade, aba in mapa_abas_unidade.items():
                        df_raw = xls[aba]
                        header_row = None
                        for r in range(min(len(df_raw), 140)):
                            for c in range(df_raw.shape[1]):
                                v = df_raw.iat[r,c]
                                if v and "nome do projeto" in _norm(str(v)):
                                    header_row = r; break
                            if header_row is not None: break
                        if header_row is None:
                            pendentes_sem_header.append((unidade, aba, df_raw))
                            continue

                        colunas = [str(v) if v is not None else f"col{c}"
                                  for c,v in enumerate(df_raw.iloc[header_row])]
                        df = df_raw.iloc[header_row+1:].copy()
                        df.columns = colunas

                        mapa_local = {}; campos_idx = {}
                        for chave,label,obrig,aliases in CAMPOS:
                            if chave == "unidade": continue
                            melhor = _melhor_coluna(colunas, aliases)
                            mapa_local[chave] = melhor
                            if melhor: campos_idx[chave] = colunas.index(melhor)

                        # Colunas de mês (Jan-Dez) — identificadas pelas células de
                        # data no próprio cabeçalho, não por nome fixo de coluna
                        # (a posição varia de aba pra aba).
                        col_mes = {}
                        for c in range(df_raw.shape[1]):
                            v = df_raw.iat[header_row, c]
                            if isinstance(v,(pd.Timestamp,datetime)):
                                col_mes[c] = (v.year, v.month)

                        mapas_cache[df_raw.shape[1]] = {"campos_idx":campos_idx, "mes_idx":col_mes}
                        _processar_aba(unidade, aba, mapa_local, col_mes, df, header_row+1)

                    # Passo 2: abas SEM cabeçalho de texto (ex: perdeu a linha de
                    # título) — reaproveita o mapeamento de outra aba com o
                    # mesmo número de colunas, assumindo dados logo após os
                    # bignumbers (linha 3 da planilha).
                    for unidade, aba, df_raw in pendentes_sem_header:
                        n_cols = df_raw.shape[1]
                        cache = mapas_cache.get(n_cols)
                        if not cache:
                            st.warning(f"⚠️ Não consegui mapear a aba **{aba}** — não achei cabeçalho "
                                      f"nem outra aba parecida (mesma qtd. de colunas) pra usar de referência.")
                            continue
                        colunas = [f"col{c}" for c in range(n_cols)]
                        mapa_local = {chave: f"col{idx}" for chave,idx in cache["campos_idx"].items()}
                        col_mes = cache["mes_idx"]
                        df = df_raw.iloc[2:].copy()
                        df.columns = colunas
                        _processar_aba(unidade, aba, mapa_local, col_mes, df, 2)

                    # Metas — aba "Parâmetros", tabela "METAS POR UNIDADE"
                    aba_param = next((a for a in nomes_abas if "parametro" in _norm(a)), None)
                    if aba_param:
                        dfp = xls[aba_param]
                        for r in range(len(dfp)):
                            nome_u = dfp.iat[r,0]
                            if nome_u is None: continue
                            nome_u_limpo = str(nome_u).strip().rstrip(":")
                            unid = normalizar_valor_lista(nome_u_limpo, unidades_disp)
                            if not unid: continue
                            # TOTAL costuma ser a última coluna numérica preenchida da linha
                            valor = None
                            for c in range(dfp.shape[1]-1, 0, -1):
                                v = dfp.iat[r,c]
                                if isinstance(v,(int,float)) and v>0:
                                    valor = v; break
                            if valor:
                                metas_linhas.append({"unidade":unid,"ano":datetime.now().year,"valor":float(valor)})

                # ── MODO 2: mapeamento manual (fallback p/ outros formatos) ────
                with st.expander("🔧 Ou importar de outro formato (mapeamento manual de 1 aba)",
                                 expanded=not modo_delga):
                    st.caption("Use isso se sua planilha não é a de Controle de Indicadores da Delga, "
                              "ou se quiser importar só uma aba específica.")
                    st.write("Abas encontradas no arquivo:")
                    st.code(", ".join(nomes_abas), language=None)

                    def _melhor_chute_aba(padrao):
                        for k in nomes_abas:
                            if padrao in _norm(k): return k
                        return nomes_abas[0]

                    c_ab1, c_ab2 = st.columns(2)
                    with c_ab1:
                        aba_proj = st.selectbox("Qual aba tem os projetos?", nomes_abas,
                            index=nomes_abas.index(_melhor_chute_aba("projeto")), key="imp_aba_proj")
                    with c_ab2:
                        opc_meta = ["— Nenhuma —"] + nomes_abas
                        chute_meta = _melhor_chute_aba("meta")
                        idx_meta = opc_meta.index(chute_meta) if any("meta" in _norm(k) for k in nomes_abas) else 0
                        aba_meta = st.selectbox("Qual aba tem as metas? (opcional)", opc_meta,
                            index=idx_meta, key="imp_aba_meta")

                    df_raw_m = xls[aba_proj]
                    header_row_m = st.number_input(
                        "Linha do cabeçalho nessa aba (1 = primeira linha)",
                        min_value=1, max_value=len(df_raw_m), value=1, key="imp_header_row") - 1
                    colunas_m = [str(v) if v is not None else f"col{c}"
                                for c,v in enumerate(df_raw_m.iloc[header_row_m])]
                    sheet_proj = df_raw_m.iloc[header_row_m+1:].copy()
                    sheet_proj.columns = colunas_m
                    st.write(f'Colunas em "{aba_proj}":')
                    st.code(", ".join(colunas_m), language=None)

                    cols_disp = ["— Não usar —"] + colunas_m
                    mapa = {}
                    cc1, cc2 = st.columns(2)
                    for i,(chave,label,obrig,aliases) in enumerate(CAMPOS):
                        chute = _melhor_coluna(colunas_m, aliases)
                        idx = cols_disp.index(chute) if chute in cols_disp else 0
                        with (cc1 if i%2==0 else cc2):
                            escolha = st.selectbox(f"{label}{' *' if obrig else ''}", cols_disp,
                                index=idx, key=f"map_{chave}")
                        mapa[chave] = None if escolha=="— Não usar —" else escolha

                    faltando = [label for chave,label,obrig,_ in CAMPOS if obrig and not mapa.get(chave)]
                    if faltando:
                        st.error(f"Aponte uma coluna pra: {', '.join(faltando)} — são obrigatórios pra importar.")
                    elif st.button("➕ Adicionar projetos dessa aba à importação", use_container_width=True):
                        for i,(_,row) in enumerate(sheet_proj.iterrows()):
                            linha = _linha_de(row, mapa, header_row_m+2+i)
                            if linha is None: continue
                            linha["_aba"] = aba_proj
                            if "_erro" in linha: linhas_erro.append(linha)
                            else: linhas_ok.append(linha)
                        st.session_state["imp_manual_ok"] = linhas_ok
                        st.session_state["imp_manual_erro"] = linhas_erro
                        st.rerun()

                    if aba_meta != "— Nenhuma —":
                        dfm_raw = xls[aba_meta]
                        colm = [str(v) if v is not None else f"col{c}"
                               for c,v in enumerate(dfm_raw.iloc[0])]
                        dfm = dfm_raw.iloc[1:].copy(); dfm.columns = colm
                        col_map2 = {_norm(c):c for c in dfm.columns}
                        for idx,row in dfm.iterrows():
                            u = row.get(col_map2.get("unidade",""))
                            if u is None or (hasattr(pd,"isna") and pd.isna(u)) or not str(u).strip(): continue
                            unid = normalizar_valor_lista(u, unidades_disp)
                            ano  = row.get(col_map2.get("ano",""))
                            val  = row.get(col_map2.get("valor da meta",""))
                            if unid and ano and val:
                                try: metas_linhas.append({"unidade":unid,"ano":int(ano),"valor":float(val)})
                                except: pass

                # Junta resultado do modo manual (se usado) com o automático
                linhas_ok = linhas_ok + st.session_state.get("imp_manual_ok", [])
                linhas_erro = linhas_erro + st.session_state.get("imp_manual_erro", [])

                st.markdown("---")
                st.markdown(f"**Prévia consolidada:** {len(linhas_ok)} projeto(s) prontos para importar · "
                           f"{len(linhas_erro)} com erro · {len(metas_linhas)} meta(s) na planilha")

                if linhas_ok:
                    with st.expander(f"Ver {len(linhas_ok)} projeto(s) prontos", expanded=len(linhas_ok)<=30):
                        prev_rows = "".join(f"""<tr>
                          <td style="font-size:11px;">{l.get('_aba','')}</td>
                          <td style="font-size:11px;font-weight:600;">{l['nome'][:45]}</td>
                          <td style="font-size:11px;">{l['unidade']}</td>
                          <td style="font-size:11px;">{l['tipo']}</td>
                          <td style="font-size:11px;text-align:right;">R$ {l['previsto_unidade']:,.0f}</td>
                          <td style="font-size:11px;">{l['mes_primeiro_retorno'] or '—'}</td>
                          <td style="font-size:11px;">{l['validacao']}</td>
                        </tr>""" for l in linhas_ok)
                        st.markdown(f"""<table class="dt"><thead><tr>
                          <th>Aba</th><th>Nome</th><th>Unidade</th><th>Tipo</th>
                          <th style="text-align:right;">Previsto</th><th>1º Retorno</th><th>Validação</th>
                        </tr></thead><tbody>{prev_rows}</tbody></table>""", unsafe_allow_html=True)

                if linhas_erro:
                    with st.expander(f"⚠️ {len(linhas_erro)} linha(s) com erro — não serão importadas"):
                        err_rows = "".join(f"""<tr>
                          <td style="font-size:11px;">{l.get('_aba','')}</td>
                          <td style="font-size:11px;">{l['linha_planilha']}</td>
                          <td style="font-size:11px;">{l['nome'][:40]}</td>
                          <td style="font-size:11px;color:#D93B3B;">{l['_erro']}</td>
                        </tr>""" for l in linhas_erro)
                        st.markdown(f"""<table class="dt"><thead><tr>
                          <th>Aba</th><th>Linha</th><th>Nome</th><th>Erro</th>
                        </tr></thead><tbody>{err_rows}</tbody></table>""", unsafe_allow_html=True)

                if metas_linhas:
                    with st.expander(f"🎯 {len(metas_linhas)} meta(s) encontrada(s)"):
                        mrows = "".join(f"""<tr>
                          <td style="font-size:11px;font-weight:600;">{m['unidade']}</td>
                          <td style="font-size:11px;">{m['ano']}</td>
                          <td style="font-size:11px;text-align:right;">R$ {m['valor']:,.0f}</td>
                        </tr>""" for m in metas_linhas)
                        st.markdown(f"""<table class="dt"><thead><tr>
                          <th>Unidade</th><th>Ano</th><th style="text-align:right;">Meta</th>
                        </tr></thead><tbody>{mrows}</tbody></table>""", unsafe_allow_html=True)

                if linhas_ok or metas_linhas:
                    total_meses_real = sum(len(l.get("_real_mensal") or {}) for l in linhas_ok)
                    st.markdown("---")
                    confirmar_imp = st.checkbox(
                        f"Confirmo a importação de {len(linhas_ok)} projeto(s), "
                        f"{len(metas_linhas)} meta(s) e {total_meses_real} lançamento(s) de real mensal.")
                    if st.button("📥 Confirmar Importação", disabled=not confirmar_imp,
                                type="primary", use_container_width=True):
                        sucesso, erros_imp = importar_projetos_lote(linhas_ok, user["id"])
                        for m in metas_linhas: set_meta(m["unidade"], m["ano"], m["valor"])
                        n_real = 0
                        for s in sucesso:
                            linha_orig = linhas_ok[s["linha"]-1]
                            for (ano_r,mes_r),val_r in (linha_orig.get("_real_mensal") or {}).items():
                                lancar_real(s["id"], ano_r, mes_r, val_r,
                                           "[Importado via Excel]", user["id"])
                                n_real += 1
                        st.success(f"✅ {len(sucesso)} projeto(s) importado(s), "
                                  f"{len(metas_linhas)} meta(s) atualizada(s) e "
                                  f"{n_real} lançamento(s) de real gravado(s).")
                        if erros_imp:
                            st.error(f"{len(erros_imp)} falharam na gravação: " +
                                    "; ".join(f"{e['nome']} ({e['erro']})" for e in erros_imp))
                        for k in ("imp_manual_ok","imp_manual_erro"):
                            st.session_state.pop(k, None)
                        st.rerun()

    with tab_reset:
        n_proj = len(listar_projetos(incluir_campeao=True))
        st.markdown(f"**🗑️ Resetar Projetos de Teste**")
        st.warning(
            f"Isso apaga **todos os {n_proj} projeto(s)** cadastrados — junto com seus "
            f"links e lançamentos de real. **Usuários, unidades e metas são mantidos.**\n\n"
            f"Use isso para limpar os dados de teste antes de começar a operação real.")
        confirmar = st.checkbox("Entendo que essa ação é irreversível e quero apagar todos os projetos.")
        if st.button("🗑️ Apagar todos os projetos", disabled=not confirmar,
                     use_container_width=True, type="primary"):
            resetar_projetos_teste()
            st.success("✅ Todos os projetos, links e lançamentos foram apagados. Usuários, unidades e metas foram mantidos.")
            st.rerun()
