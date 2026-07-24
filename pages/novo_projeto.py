import streamlit as st
import base64
import pandas as pd
from datetime import date, datetime, timedelta
import plotly.graph_objects as go
from database import (listar_unidades, criar_projeto, add_link,
                      TIPOS_PROJETO, VA_GGF_OPTS, STATUS_OPTS,
                      EXTRA_DRE_TIPOS, CAMPOS_A3, get_a3, salvar_a3,
                      add_a3_midia, get_a3_midias, del_a3_midia,
                      add_evidencia, get_evidencias, del_evidencia,
                      listar_atividades, add_atividade, del_atividade,
                      progresso_plan_atividade, atividade_atual,
                      get_projeto, atualizar_projeto)

DATA_FIM_PROJETOS_APLICADOS = date(2027, 1, 1)
COLS_ESTRUTURA = ["Atividade","Responsável","Ação","Início Previsto","Término Previsto","% Progresso Real"]

def pode_editar(user, unidade_nome):
    if user["perfil"] == "admin": return True
    if user["perfil"] in ("facilitador","gestor"):
        return user.get("unidade") == unidade_nome
    return False

def _parse_date(v):
    if not v: return None
    if isinstance(v, datetime): return v.date()
    if isinstance(v, date): return v
    try: return datetime.strptime(str(v)[:10], "%Y-%m-%d").date()
    except Exception: return None

def _arquivo_para_b64(uploaded_file, limite_mb=8):
    if not uploaded_file: return None, None
    dados = uploaded_file.read()
    if len(dados) > limite_mb * 1024 * 1024:
        return None, f"Arquivo maior que {limite_mb}MB — deixa o backup muito pesado, escolhe um menor."
    return base64.b64encode(dados).decode("utf-8"), None

def _extrair_imagens_a3(ws):
    """Localiza as imagens coladas na aba 'Escopo A3' e descobre em qual dos
    6 blocos cada uma está, pela posição da âncora na planilha (mesma área
    de células que o texto daquele bloco ocupa)."""
    blocos = {  # campo -> (col0, col1, row0, row1), 0-indexed
        "objetivo_geral": (2, 7, 14, 19),
        "proposta_desenvolvimento": (9, 15, 14, 42),
        "situacao_atual": (2, 7, 22, 34),
        "metas_entregas": (2, 7, 37, 42),
        "premissas_restricoes": (2, 7, 45, 55),
        "acompanhamento_indicadores": (9, 15, 45, 55),
    }
    resultado = {k: [] for k in blocos}
    for i, img in enumerate(getattr(ws, "_images", [])):
        try:
            col0, row0 = img.anchor._from.col, img.anchor._from.row
        except Exception:
            continue
        alvo = next((campo for campo, (c0,c1,r0,r1) in blocos.items()
                     if c0 <= col0 <= c1 and r0 <= row0 <= r1), None)
        if not alvo: continue
        try:
            dados_bytes = img._data()
        except Exception:
            continue
        ext = (getattr(img, "format", None) or "png").lower()
        resultado[alvo].append((f"imagem_{alvo}_{i+1}.{ext}", f"image/{ext}", dados_bytes))
    return resultado

def _importar_gestao_projetos_excel(arquivo):
    """Lê o Excel padrão 'Gestão de Projetos' (abas 'Escopo A3' e
    'Estrutura', modelo que a Delga já usa) e devolve um dict pronto pra
    pré-popular a criação do projeto: cabeçalho, os 6 blocos do A3 (texto
    + imagens coladas) e a lista de atividades. Se alguma aba não existir
    ou algum campo não bater, simplesmente vem vazio — nunca derruba o
    resto do import."""
    import openpyxl
    wb = openpyxl.load_workbook(arquivo, data_only=True)
    resultado = {"cabecalho": {}, "a3": {}, "imagens": {}, "atividades": []}

    if "Escopo A3" in wb.sheetnames:
        ws = wb["Escopo A3"]
        def val(coord):
            try:
                v = ws[coord].value
                return str(v).strip() if v is not None else ""
            except Exception:
                return ""
        resultado["cabecalho"] = {
            "nome": val("D2"),
            "numero_projeto": val("P5"),
            "integrantes": val("G5"),
            "lider_projeto": val("D7"),
            "unidade": val("G7"),
            "revisao": val("P2").replace("Rev.:", "").replace("Rev:", "").strip(),
        }
        resultado["a3"] = {
            "objetivo_geral": val("C15"),
            "proposta_desenvolvimento": val("J15"),
            "situacao_atual": val("C23"),
            "metas_entregas": val("C38"),
            "premissas_restricoes": val("C46"),
            "acompanhamento_indicadores": val("J46"),
        }
        try:
            resultado["imagens"] = _extrair_imagens_a3(ws)
        except Exception:
            resultado["imagens"] = {}

    if "Estrutura" in wb.sheetnames:
        ws = wb["Estrutura"]
        header_row, col_map = None, {}
        for r in range(1, 6):
            for c in range(1, 12):
                v = ws.cell(r, c).value
                if v and "atividade" in str(v).strip().lower():
                    header_row = r; break
            if header_row: break
        if header_row:
            for c in range(1, 12):
                v = ws.cell(header_row, c).value
                if not v: continue
                vn = str(v).strip().lower().replace("\n", " ")
                if "atividade" in vn: col_map["atividade"] = c
                elif "responsav" in vn: col_map["responsavel"] = c
                elif vn.strip() == "ação" or "acao" in vn: col_map["acao"] = c
                elif "inicio previsto" in vn or "início previsto" in vn: col_map["inicio"] = c
                elif "termino previsto" in vn or "término previsto" in vn: col_map["termino"] = c
                elif "progresso real" in vn: col_map["progresso_real"] = c
            if "atividade" in col_map:
                for r in range(header_row+1, header_row+300):
                    nome = ws.cell(r, col_map["atividade"]).value
                    if not nome or not str(nome).strip(): continue
                    def g(chave):
                        c = col_map.get(chave)
                        return ws.cell(r, c).value if c else None
                    resultado["atividades"].append({
                        "nome": str(nome).strip(),
                        "responsavel": str(g("responsavel") or ""),
                        "acao": str(g("acao") or ""),
                        "inicio": g("inicio"),
                        "termino": g("termino"),
                        "progresso_real": g("progresso_real") or 0,
                    })
    return resultado

def _estrutura_df_vazia(n=10):
    return pd.DataFrame({
        "Atividade": [""]*n, "Responsável": [""]*n, "Ação": [""]*n,
        "Início Previsto": [None]*n, "Término Previsto": [None]*n,
        "% Progresso Real": [0]*n,
    })

def _estrutura_col_config():
    return {
        "Início Previsto": st.column_config.DateColumn("Início Previsto", format="DD/MM/YYYY"),
        "Término Previsto": st.column_config.DateColumn("Término Previsto", format="DD/MM/YYYY"),
        "% Progresso Real": st.column_config.NumberColumn("% Progresso Real", min_value=0, max_value=100, step=5),
    }

def _linhas_validas(df):
    """Filtra só as linhas com Atividade preenchida (ignora as vazias da grade)."""
    return [row for _, row in df.iterrows() if str(row.get("Atividade") or "").strip()]

def _resumo_progresso_plan(linhas):
    if not linhas: return
    partes = []
    for row in linhas:
        p = progresso_plan_atividade(row.get("Início Previsto"), row.get("Término Previsto"))
        if p is not None:
            partes.append(f"{str(row['Atividade']).strip()}: {p:.0f}%")
    if partes:
        st.caption("📐 **% Progresso Plan** (calculado sozinho, comparando hoje com as datas previstas): " +
                  " · ".join(partes))

# =====================================================================
# GANTT — calculado em cima da Estrutura, não é preenchido à parte
# =====================================================================
def build_gantt(atividades, colors):
    """atividades: lista de dicts com nome/inicio_previsto/termino_previsto/
    progresso_real (aceita tanto registros do banco quanto linhas da grade
    em memória, desde que usem essas chaves). Visual em pílulas coloridas,
    uma cor por atividade, nome dentro da barra logo no início, e a escala
    do eixo (dia/semana/mês) se ajusta sozinha ao tamanho do projeto."""
    NAVY=colors.get("NAVY","#0B0F2B"); RED=colors.get("RED","#D93B3B")
    SILVER=colors.get("SILVER","#8A9BB0")

    linhas = []
    for a in atividades:
        ini = _parse_date(a.get("inicio_previsto")); fim = _parse_date(a.get("termino_previsto"))
        if not ini or not fim or fim < ini: continue
        linhas.append({"nome": a.get("nome") or "(sem nome)", "ini": ini, "fim": fim,
                       "prog": max(0.0, min(1.0, (a.get("progresso_real") or 0)/100))})
    if not linhas:
        return None

    n = len(linhas)
    tarefas = [f"#{i+1}  {l['nome']}" for i, l in enumerate(linhas)]
    data_min = min(l["ini"] for l in linhas)
    data_max = max(l["fim"] for l in linhas)
    hoje = date.today()
    span_atividades = (data_max - data_min).days
    folga = max(2, round(span_atividades * 0.05))
    janela_ini = data_min - timedelta(days=folga)
    janela_fim = data_max + timedelta(days=folga)
    # só estica a janela pra incluir "hoje" se ele estiver relativamente perto
    # do período das atividades — senão um projeto de poucos dias lá no
    # futuro/passado fica esmagado numa janela gigante só pra caber a
    # linha de "hoje". Limite: até 1x o próprio tamanho do projeto de
    # distância (mínimo 30 dias de tolerância).
    tolerancia = timedelta(days=max(30, span_atividades))
    mostra_linha_hoje = (janela_ini - tolerancia) <= hoje <= (janela_fim + tolerancia)
    if mostra_linha_hoje:
        janela_ini = min(janela_ini, hoje - timedelta(days=folga))
        janela_fim = max(janela_fim, hoje + timedelta(days=folga))
    total_dias_janela = (janela_fim - janela_ini).days

    fig = go.Figure()

    # fins de semana sombreados (some sozinho se o projeto for longo demais
    # pra fazer sentido visual)
    if total_dias_janela <= 220:
        cursor = janela_ini
        while cursor <= janela_fim:
            if cursor.weekday() >= 5:
                fig.add_vrect(x0=datetime.combine(cursor, datetime.min.time()),
                              x1=datetime.combine(cursor+timedelta(days=1), datetime.min.time()),
                              fillcolor="#F5F6FA", line_width=0, layer="below")
            cursor += timedelta(days=1)

    # separador de mês
    cursor = date(janela_ini.year, janela_ini.month, 1)
    while cursor <= janela_fim:
        if cursor >= janela_ini:
            fig.add_vline(x=datetime.combine(cursor, datetime.min.time()),
                         line_width=1, line_color="#E4E7EF")
        cursor = date(cursor.year+1, 1, 1) if cursor.month == 12 else date(cursor.year, cursor.month+1, 1)

    # barras estilo grade: trilho fino com borda (planejado) + preenchimento
    # sólido conforme % Real, cor por status (não por atividade) — igual ao
    # modelo escolhido
    espessura = 0.55
    for i, l in enumerate(linhas):
        nome, ini, fim, prog = l["nome"], l["ini"], l["fim"], l["prog"]
        total_dias = (fim-ini).days + 1
        atrasada = prog < 1 and fim < hoje
        concluida = prog >= 1
        cor = RED if atrasada else ("#639922" if concluida else ("#378ADD" if prog > 0 else SILVER))

        fig.add_trace(go.Bar(
            x=[total_dias], y=[tarefas[i]], base=[ini], orientation="h",
            marker=dict(color="white", cornerradius=4, line=dict(color=cor, width=1.2)),
            showlegend=False, hoverinfo="skip", width=espessura))
        dias_preenchidos = round(total_dias*prog)
        if dias_preenchidos > 0:
            fig.add_trace(go.Bar(
                x=[dias_preenchidos], y=[tarefas[i]], base=[ini], orientation="h",
                marker=dict(color=cor, cornerradius=3), showlegend=False,
                hovertemplate=f"<b>{nome}</b><br>{ini:%d/%m/%y} – {fim:%d/%m/%y} ({total_dias}d)<br>Real: {prog*100:.0f}%<extra></extra>",
                width=espessura*0.82))

    # linha de hoje — só desenha se ela cabe na janela sem distorcer o resto
    if mostra_linha_hoje:
        fig.add_vline(x=datetime.combine(hoje, datetime.min.time()), line_width=2, line_color=RED,
                     annotation_text="Hoje", annotation_position="top", annotation_font=dict(size=10, color=RED))


    # granularidade do eixo se ajusta ao tamanho total do período
    if total_dias_janela <= 21: dtick, fmt = "D1", "%d/%m"
    elif total_dias_janela <= 70: dtick, fmt = 86400000*2, "%d/%m"
    elif total_dias_janela <= 220: dtick, fmt = "D7", "%d/%m"
    elif total_dias_janela <= 730: dtick, fmt = "M1", "%b/%y"
    else: dtick, fmt = "M3", "%b/%y"

    fig.update_xaxes(type="date", range=[datetime.combine(janela_ini, datetime.min.time()),
                                          datetime.combine(janela_fim, datetime.min.time())],
                     dtick=dtick, tickformat=fmt, tickangle=0, gridcolor="#EEF0F5",
                     side="top", showgrid=True)
    fig.update_yaxes(autorange="reversed", title=None, showgrid=False)
    fig.update_layout(
        barmode="overlay", height=max(240, 46*n + 60),
        margin=dict(l=10, r=20, t=40, b=10),
        bargap=0.35,
        paper_bgcolor="white", plot_bgcolor="white",
        font=dict(family="Inter", size=12, color=NAVY),
        showlegend=False)
    return fig

def _linhas_para_gantt(linhas_df):
    return [{"nome": r["Atividade"], "inicio_previsto": r.get("Início Previsto"),
             "termino_previsto": r.get("Término Previsto"), "progresso_real": r.get("% Progresso Real") or 0}
            for r in linhas_df]

# =====================================================================
# ESTRUTURA — grade fluida estilo Excel (usada tanto na criação, em
# memória, quanto na edição de um projeto já existente, ligada ao banco)
# =====================================================================
def _render_estrutura_db(pid, user, colors):
    """Versão ligada ao banco — usada ao editar um projeto Novo Projeto já
    criado (via Minha Unidade)."""
    ativs_db = listar_atividades(pid)
    terminos_antes = sorted(str(a.get("termino_previsto") or "") for a in ativs_db)

    if ativs_db:
        df = pd.DataFrame([{
            "Atividade": a.get("nome") or "", "Responsável": a.get("responsavel") or "",
            "Ação": a.get("acao") or "", "Início Previsto": _parse_date(a.get("inicio_previsto")),
            "Término Previsto": _parse_date(a.get("termino_previsto")),
            "% Progresso Real": a.get("progresso_real") or 0,
        } for a in ativs_db])
        df = pd.concat([df, _estrutura_df_vazia(3)], ignore_index=True)
    else:
        df = _estrutura_df_vazia(10)

    st.caption("Grade estilo planilha — edite direto nas células, arraste pra baixo pra preencher mais rápido. "
              "Use o **+** no fim da tabela pra mais linhas.")
    edited = st.data_editor(df[COLS_ESTRUTURA], key=f"est_editor_{pid}", num_rows="dynamic",
                             use_container_width=True, hide_index=True, column_config=_estrutura_col_config())

    linhas = _linhas_validas(edited)
    _resumo_progresso_plan(linhas)

    if st.button("💾 Salvar Estrutura", key=f"salvar_est_{pid}", type="primary", use_container_width=True):
        terminos_depois = sorted(str(_parse_date(row.get("Término Previsto")) or "") for row in linhas)
        for a in ativs_db:
            del_atividade(a["id"])
        for row in linhas:
            add_atividade(pid, {
                "nome": str(row["Atividade"]).strip(),
                "responsavel": str(row.get("Responsável") or ""),
                "acao": str(row.get("Ação") or ""),
                "inicio_previsto": str(_parse_date(row.get("Início Previsto")) or "") or None,
                "termino_previsto": str(_parse_date(row.get("Término Previsto")) or "") or None,
                "progresso_real": float(row.get("% Progresso Real") or 0),
            })
        if ativs_db and terminos_antes != terminos_depois:
            p_atual = get_projeto(pid)
            atualizar_projeto(pid, {"replanejamentos": (p_atual.get("replanejamentos") or 0) + 1}, user["id"])
        st.success("✅ Estrutura salva!"); st.rerun()

    atual = atividade_atual(pid)
    if atual:
        st.info(f"📍 **Atual Atribuição (automática):** {atual['nome']} · "
               f"Responsável: {atual.get('responsavel') or '—'} · "
               f"Término Previsto: {str(atual.get('termino_previsto') or '—')[:10]}")
    elif ativs_db:
        st.success("✅ Todas as atividades concluídas!")

def _hoje_fora_do_periodo(atividades):
    """True se 'hoje' estiver longe demais do período das atividades pro
    Gantt ter incluído a linha de hoje (mesmo critério do build_gantt)."""
    datas_ini = [_parse_date(a.get("inicio_previsto")) for a in atividades]
    datas_fim = [_parse_date(a.get("termino_previsto")) for a in atividades]
    datas_ini = [d for d in datas_ini if d]; datas_fim = [d for d in datas_fim if d]
    if not datas_ini or not datas_fim: return False, None
    data_min, data_max = min(datas_ini), max(datas_fim)
    hoje = date.today()
    span = (data_max - data_min).days
    tolerancia = timedelta(days=max(30, span))
    return not ((data_min - tolerancia) <= hoje <= (data_max + tolerancia)), hoje

def _render_gantt_db(pid, colors):
    ativs = [a for a in listar_atividades(pid) if a.get("inicio_previsto") and a.get("termino_previsto")]
    if not ativs:
        st.info("Preencha a **Estrutura** com Início e Término previstos pra ver o Gantt aqui.")
        return
    st.caption("A barra tem o tamanho do planejado e é pintada conforme o % Progresso Real. "
              "Vermelho na barra = atividade atrasada.")
    fora, hoje = _hoje_fora_do_periodo(ativs)
    if fora and hoje:
        st.caption(f"📅 Hoje ({hoje:%d/%m/%Y}) está fora do período das atividades — "
                  f"a linha de hoje não aparece pra não distorcer a escala do gráfico.")
    fig = build_gantt(ativs, colors)
    if fig:
        st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})

# =====================================================================
# A3 — 6 blocos + imagens inline + evidências gerais (versão ligada ao
# banco, usada ao editar um projeto Novo Projeto já criado)
# =====================================================================
def _render_a3_db(pid, colors):
    a3 = get_a3(pid)
    st.caption("Preencha cada bloco do A3. Pode anexar imagem em qualquer bloco (fica visível ali dentro — "
              "ex: um fluxo na Proposta/Desenvolvimento). Anexos que não são imagem (PPT, Excel, PDF) vão "
              "no espaço de Evidências, no fim da página.")
    valores = {}
    for campo, label in CAMPOS_A3:
        with st.expander(f"**{label}**", expanded=True):
            valores[campo] = st.text_area(label, value=a3.get(campo) or "", height=130,
                                           key=f"a3_{campo}_{pid}", label_visibility="collapsed")
            imgs = get_a3_midias(pid, campo)
            if imgs:
                cols = st.columns(min(len(imgs), 4))
                for i, img in enumerate(imgs):
                    with cols[i % 4]:
                        try:
                            st.image(base64.b64decode(img["dados_b64"]), caption=img.get("nome_arquivo"),
                                     use_container_width=True)
                        except Exception:
                            st.caption(f"📎 {img.get('nome_arquivo')}")
                        if st.button("🗑️ remover", key=f"delimg_{img['id']}", use_container_width=True):
                            del_a3_midia(img["id"]); st.rerun()
            # a key muda a cada upload processado — assim o file_uploader "esquece"
            # o arquivo anterior e não fica reprocessando/duplicando a cada rerun
            ctr_key = f"upimg_ctr_{campo}_{pid}"
            ctr = st.session_state.get(ctr_key, 0)
            up = st.file_uploader(f"Adicionar imagem — {label}", type=["png","jpg","jpeg","gif","webp"],
                                   key=f"upimg_{campo}_{pid}_{ctr}", label_visibility="collapsed")
            if up is not None:
                b64, erro = _arquivo_para_b64(up)
                if erro:
                    st.error(erro)
                elif b64:
                    add_a3_midia(pid, campo, up.name, up.type or "image/png", b64)
                    st.session_state[ctr_key] = ctr + 1
                    st.success("✅ Imagem adicionada."); st.rerun()

    if st.button("💾 Salvar A3", type="primary", use_container_width=True, key=f"salvar_a3_{pid}"):
        salvar_a3(pid, valores)
        st.success("✅ A3 salvo!")

    st.markdown("---")
    st.markdown("**📎 Evidências / Anexos** *(PPT, Excel, PDF ou qualquer arquivo de apoio)*")
    evids = get_evidencias(pid)
    for e in evids:
        c1,c2,c3 = st.columns([5,2,1])
        with c1: st.markdown(f"📄 **{e.get('nome_arquivo')}**")
        with c2: st.caption(e.get("mime_type") or "")
        with c3:
            if st.button("🗑️", key=f"delev_{e['id']}"): del_evidencia(e["id"]); st.rerun()
    ctr_ev_key = f"upev_ctr_{pid}"
    ctr_ev = st.session_state.get(ctr_ev_key, 0)
    up_ev = st.file_uploader("Adicionar evidência", key=f"upev_{pid}_{ctr_ev}",
                              type=["pdf","ppt","pptx","xls","xlsx","doc","docx","png","jpg","jpeg"])
    if up_ev is not None:
        b64, erro = _arquivo_para_b64(up_ev)
        if erro:
            st.error(erro)
        elif b64:
            add_evidencia(pid, up_ev.name, up_ev.type or "application/octet-stream", b64)
            st.session_state[ctr_ev_key] = ctr_ev + 1
            st.success("✅ Evidência anexada."); st.rerun()

def _processar_arquivo_pronto():
    """Lê o arquivo em st.session_state['npn_arq_pronto'] (se tiver e ainda
    não tiver sido processado) e pré-popula os campos de Fundamentos/A3/
    Estrutura/imagens. PRECISA rodar antes de qualquer widget desses campos
    ser desenhado na tela — o Streamlit não deixa mudar o valor de um campo
    depois que ele já apareceu na mesma rodada."""
    if not str(st.session_state.get("npn_tem_arquivo") or "").startswith("Sim"):
        return None
    arq_pronto = st.session_state.get("npn_arq_pronto")
    if arq_pronto is None or st.session_state.get("npn_arq_processado"):
        return None
    try:
        dados_imp = _importar_gestao_projetos_excel(arq_pronto)
    except Exception as e:
        return ("error", f"Não consegui ler esse arquivo: {e}")

    cab = dados_imp["cabecalho"]
    if cab.get("nome"): st.session_state["npn_nome"] = cab["nome"]
    if cab.get("numero_projeto"): st.session_state["npn_numero"] = cab["numero_projeto"]
    if cab.get("lider_projeto"): st.session_state["npn_lider"] = cab["lider_projeto"]
    if cab.get("integrantes"): st.session_state["npn_integrantes"] = cab["integrantes"]
    if cab.get("revisao"): st.session_state["npn_revisao"] = cab["revisao"]
    if cab.get("unidade"):
        alvo = cab["unidade"].strip().lower()
        for u in listar_unidades():
            if u["nome"].strip().lower() == alvo:
                st.session_state["npn_uni"] = u["nome"]; break

    for campo, texto in dados_imp["a3"].items():
        if texto: st.session_state[f"npn_a3_{campo}"] = texto

    n_imgs = 0
    for campo, imgs in dados_imp.get("imagens", {}).items():
        if not imgs: continue
        chave = f"npn_imgs_{campo}"
        if chave not in st.session_state: st.session_state[chave] = []
        for nome_img, mime_img, dados_bytes in imgs:
            st.session_state[chave].append({
                "nome": nome_img, "tipo": mime_img,
                "b64": base64.b64encode(dados_bytes).decode("utf-8")})
            n_imgs += 1

    if dados_imp["atividades"]:
        df_imp = pd.DataFrame([{
            "Atividade": a["nome"], "Responsável": a.get("responsavel") or "",
            "Ação": a.get("acao") or "",
            "Início Previsto": _parse_date(a.get("inicio")),
            "Término Previsto": _parse_date(a.get("termino")),
            "% Progresso Real": a.get("progresso_real") or 0,
        } for a in dados_imp["atividades"]])
        df_imp = pd.concat([df_imp, _estrutura_df_vazia(3)], ignore_index=True)
        st.session_state["npn_estrutura_df"] = df_imp
        st.session_state.pop("npn_estrutura_editor", None)  # força a grade a reler a seed nova

    st.session_state["npn_arq_processado"] = True
    n_ativ_imp = len(dados_imp["atividades"])
    n_a3_imp = sum(1 for v in dados_imp["a3"].values() if v)
    return ("success", f"✅ Importei {n_ativ_imp} atividade(s), {n_a3_imp} bloco(s) do A3 e "
                       f"{n_imgs} imagem(ns)! Role a página pra revisar tudo antes de clicar em Criar Projeto.")

# Aliases mantidos pra compatibilidade com quem já importa esses nomes
_render_a3 = _render_a3_db
_render_estrutura = _render_estrutura_db
_render_gantt = _render_gantt_db

# =====================================================================
# NOVO PROJETO — página única: Fundamentos > A3 > Estrutura > Datas >
# Gantt > Checklist > Status/Observações > Criar. Nada é salvo até o
# botão final — A3/Estrutura/imagens ficam em memória (session_state)
# até lá, porque o projeto ainda não existe.
# =====================================================================
def _limpar_estado_novo_projeto():
    for k in list(st.session_state.keys()):
        if k.startswith("npn_") or k.startswith("a3n_") or k.startswith("upimgn_") or k.startswith("upevn_"):
            del st.session_state[k]

def _render_novo_projeto(user, colors):
    NAVY=colors.get("NAVY","#0B0F2B"); GREEN=colors.get("GREEN","#1AA260")

    # Processa um arquivo pronto ANTES de desenhar qualquer campo — precisa
    # ser aqui em cima, senão o Streamlit barra a alteração dos campos que
    # já renderizaram nesta mesma rodada.
    resultado_import = _processar_arquivo_pronto()

    # ── 0. Primeira pergunta: já tem o arquivo pronto? ──────────────────
    st.markdown("### 👋 Antes de começar")
    st.markdown("**Você já tem o Excel do projeto (Escopo A3 + Estrutura) preenchido?**")
    st.caption("Se tiver bem preenchido, eu já trago Nome, Nº do Projeto, Líder, Integrantes, os 6 blocos "
              "do A3 (com imagens coladas!) e todas as atividades da Estrutura — você só revisa e ajusta "
              "o que quiser conforme for descendo a página. Se não tiver, sem problema, é só preencher "
              "tudo manualmente daqui pra baixo.")
    tem_arquivo = st.radio("Tem o arquivo pronto?", ["Ainda não, vou preencher manualmente","Sim, já tenho!"],
                           key="npn_tem_arquivo", horizontal=True, label_visibility="collapsed")
    if tem_arquivo.startswith("Sim"):
        st.file_uploader("Excel do projeto (Escopo A3 + Estrutura)", type=["xlsx"], key="npn_arq_pronto")
        if resultado_import:
            tipo_msg, texto_msg = resultado_import
            (st.error if tipo_msg == "error" else st.success)(texto_msg)
        if st.session_state.get("npn_arq_processado"):
            st.info("📄 Arquivo já importado — os campos daqui pra baixo já vêm pré-preenchidos. Pode "
                   "revisar e ajustar qualquer coisa normalmente antes de criar o projeto.")
            if st.button("↩️ Importar outro arquivo / desfazer", key="npn_desfazer_import"):
                st.session_state.pop("npn_arq_processado", None)
                st.session_state.pop("npn_arq_pronto", None)
                st.rerun()

    st.markdown("---")

    unidades = listar_unidades()
    nomes_u  = [u["nome"] for u in unidades]

    if user["perfil"] == "admin":
        unidade_sel = st.selectbox("Unidade:", nomes_u, key="npn_uni")
    elif user["perfil"] in ("facilitador","gestor","cost_control"):
        unidade_sel = st.selectbox("Unidade:", nomes_u, key="npn_uni")
        if user.get("unidade") and unidade_sel != user.get("unidade"):
            st.info(f"👁️ Você cria projetos apenas em **{user.get('unidade')}**")
    else:
        unidade_sel = user.get("unidade","")
        if unidade_sel not in nomes_u:
            st.warning("Unidade não configurada."); return

    if not pode_editar(user, unidade_sel):
        st.warning(f"⛔ Você só pode criar projetos em **{user.get('unidade','')}**.")
        return

    # ── 1. Fundamentos ──────────────────────────────────────────────────
    st.markdown("### 🧱 Fundamentos")
    c1,c2,c3 = st.columns(3)
    with c1: tipo = st.selectbox("Tipo *", TIPOS_PROJETO, key="npn_tipo")
    with c2: va   = st.selectbox("VA / GGF / Material Auxiliar *", VA_GGF_OPTS, key="npn_va")
    with c3: resp = st.text_input("Responsável", key="npn_resp")

    is_extra = tipo in EXTRA_DRE_TIPOS
    if is_extra:
        st.markdown(f'<div style="background:#F3E8FF;border-left:3px solid #9B59B6;border-radius:0 6px 6px 0;'
                    f'padding:8px 14px;font-size:11px;color:#6C3483;"><b>↷ Extra DRE</b> — {tipo} não gera '
                    f'lançamento de real mensal.</div>', unsafe_allow_html=True)
    else:
        st.markdown(f'<div style="background:#E6F4EC;border-left:3px solid {GREEN};border-radius:0 6px 6px 0;'
                    f'padding:8px 14px;font-size:11px;color:#1A5C2E;"><b>✓ Dentro do DRE</b> — {tipo} impacta '
                    f'diretamente o DRE.</div>', unsafe_allow_html=True)

    nome = st.text_input("Nome do Projeto *", key="npn_nome")
    desc = st.text_area("Descrição / Objetivo", height=70, key="npn_desc")

    st.markdown("**Cabeçalho do A3**")
    c1,c2,c3 = st.columns(3)
    with c1: numero_p = st.text_input("Nº do Projeto", key="npn_numero")
    with c2: lider_p = st.text_input("Líder do Projeto", key="npn_lider")
    with c3: revisao_p = st.text_input("Revisão", value="Rev. 00", key="npn_revisao")
    integrantes_p = st.text_input("Integrantes", placeholder="Nomes separados por vírgula", key="npn_integrantes")

    ganho_unico = st.checkbox(
        "🎯 Ganho Único — retorno pontual, só no mês do 1º retorno",
        help="Concentra o valor inteiro no mês de retorno, sem ratear em 12 meses.", key="npn_gu")
    previsto = st.number_input("Valor Previsto (R$) *", min_value=0.0, step=1000.0, format="%.2f", key="npn_previsto")

    st.markdown("**Links (SharePoint / OneDrive)**")
    c1,c2 = st.columns([2,4])
    with c1: link1_tit = st.text_input("Nome 1", placeholder="ex: Planilha de Apoio", key="npn_l1t")
    with c2: link1_url = st.text_input("URL 1", placeholder="https://...", key="npn_l1u")

    st.markdown("---")

    # ── 2. A3 ────────────────────────────────────────────────────────────
    st.markdown("### 📋 A3")
    st.caption("Preencha cada bloco. Pode anexar imagem em qualquer um (fica visível ali dentro — ex: um "
              "fluxo na Proposta/Desenvolvimento). Isso tudo só é gravado quando você clicar em "
              "**Criar Projeto**, no fim da página.")
    for campo, label in CAMPOS_A3:
        with st.expander(f"**{label}**", expanded=False):
            st.text_area(label, height=110, key=f"npn_a3_{campo}", label_visibility="collapsed")
            imgs_key = f"npn_imgs_{campo}"
            if imgs_key not in st.session_state:
                st.session_state[imgs_key] = []
            imgs = st.session_state[imgs_key]
            if imgs:
                cols = st.columns(min(len(imgs), 4))
                for i, img in enumerate(imgs):
                    with cols[i % 4]:
                        try:
                            st.image(base64.b64decode(img["b64"]), caption=img["nome"], use_container_width=True)
                        except Exception:
                            st.caption(f"📎 {img['nome']}")
                        if st.button("🗑️ remover", key=f"delimgn_{campo}_{i}", use_container_width=True):
                            imgs.pop(i); st.rerun()
            ctr = st.session_state.get(f"upimgn_ctr_{campo}", 0)
            up = st.file_uploader(f"Adicionar imagem — {label}", type=["png","jpg","jpeg","gif","webp"],
                                   key=f"upimgn_{campo}_{ctr}", label_visibility="collapsed")
            if up is not None:
                b64, erro = _arquivo_para_b64(up)
                if erro:
                    st.error(erro)
                elif b64:
                    imgs.append({"nome": up.name, "tipo": up.type or "image/png", "b64": b64})
                    st.session_state[f"upimgn_ctr_{campo}"] = ctr + 1
                    st.success("✅ Imagem adicionada."); st.rerun()

    st.markdown("**📎 Evidências / Anexos** *(PPT, Excel, PDF ou qualquer arquivo de apoio)*")
    if "npn_evid" not in st.session_state:
        st.session_state["npn_evid"] = []
    evid = st.session_state["npn_evid"]
    for i, e in enumerate(evid):
        c1,c2,c3 = st.columns([5,2,1])
        with c1: st.markdown(f"📄 **{e['nome']}**")
        with c2: st.caption(e.get("tipo") or "")
        with c3:
            if st.button("🗑️", key=f"delevn_{i}"): evid.pop(i); st.rerun()
    ctr_ev = st.session_state.get("upevn_ctr", 0)
    up_ev = st.file_uploader("Adicionar evidência", key=f"upevn_{ctr_ev}",
                              type=["pdf","ppt","pptx","xls","xlsx","doc","docx","png","jpg","jpeg"])
    if up_ev is not None:
        b64, erro = _arquivo_para_b64(up_ev)
        if erro:
            st.error(erro)
        elif b64:
            evid.append({"nome": up_ev.name, "tipo": up_ev.type or "application/octet-stream", "b64": b64})
            st.session_state["upevn_ctr"] = ctr_ev + 1
            st.success("✅ Evidência anexada."); st.rerun()

    st.markdown("---")

    # ── 3. Estrutura ─────────────────────────────────────────────────────
    st.markdown("### 🗓️ Estrutura")
    st.caption("Grade estilo planilha — Atividade, Responsável, Ação, Início/Término Previsto e "
              "% Progresso Real. Use o **+** no fim da tabela pra adicionar mais linhas.")
    if "npn_estrutura_df" not in st.session_state:
        st.session_state["npn_estrutura_df"] = _estrutura_df_vazia(10)
    edited_estrutura = st.data_editor(
        st.session_state["npn_estrutura_df"][COLS_ESTRUTURA], key="npn_estrutura_editor",
        num_rows="dynamic", use_container_width=True, hide_index=True,
        column_config=_estrutura_col_config())
    linhas_ativ = _linhas_validas(edited_estrutura)
    _resumo_progresso_plan(linhas_ativ)

    st.markdown("---")

    # ── 4. Datas do Projeto (calculadas a partir da Estrutura) ─────────────
    st.markdown("### 📅 Datas do Projeto")
    inicios = [_parse_date(r.get("Início Previsto")) for r in linhas_ativ]
    terminos = [_parse_date(r.get("Término Previsto")) for r in linhas_ativ]
    inicios = [d for d in inicios if d]; terminos = [d for d in terminos if d]
    data_inicio_auto = min(inicios) if inicios else None
    data_fim_auto = max(terminos) if terminos else None
    c1,c2,c3 = st.columns(3)
    with c1:
        st.text_input("Data de Início do Projeto", value=data_inicio_auto.strftime("%d/%m/%Y") if data_inicio_auto else "—",
                     disabled=True, help="Calculada sozinha: a data mais antiga entre as atividades da Estrutura.")
    with c2:
        st.text_input("Data de Fim do Projeto", value=data_fim_auto.strftime("%d/%m/%Y") if data_fim_auto else "—",
                     disabled=True, help="Calculada sozinha: a data mais futura entre as atividades da Estrutura.")
    with c3:
        mpr = st.date_input("Ganho a partir de... *", format="DD/MM/YYYY", key="npn_mpr",
            help="Mês em que o projeto começa a gerar ganho financeiro — você escolhe, não depende da Estrutura.")
    if not inicios:
        st.caption("Preencha a Estrutura acima com Início/Término Previsto pra essas datas calcularem sozinhas.")

    st.markdown("---")

    # ── 5. Gantt (preview, calculado a partir da grade acima) ──────────────
    st.markdown("### 📊 Gantt")
    if linhas_ativ:
        st.caption("Prévia — a barra tem o tamanho do planejado e é pintada conforme o % Progresso Real.")
        ativs_gantt = _linhas_para_gantt(linhas_ativ)
        fora, hoje = _hoje_fora_do_periodo(ativs_gantt)
        if fora and hoje:
            st.caption(f"📅 Hoje ({hoje:%d/%m/%Y}) está fora do período das atividades — "
                      f"a linha de hoje não aparece pra não distorcer a escala do gráfico.")
        fig = build_gantt(ativs_gantt, colors)
        if fig:
            st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})
    else:
        st.info("Preencha a Estrutura acima pra ver o Gantt aqui.")

    st.markdown("---")

    # ── 6. Checklist (por último) ────────────────────────────────────────
    st.markdown("### ✅ Checklist")
    st.caption("Quando os 3 estiverem marcados, o projeto vai para validação de Custos.")
    c1,c2,c3 = st.columns(3)
    with c1: ck_a3  = st.checkbox("A3 desenvolvido", key="npn_cka3")
    with c2: ck_mem = st.checkbox("Memória de Cálculo desenvolvida", key="npn_ckmem")
    with c3: ck_for = st.checkbox("Formalizado com Custos", key="npn_ckfor")

    st.markdown("### Status e Observações")
    status = st.selectbox("Status", STATUS_OPTS, key="npn_status")
    obs = st.text_area("Observações", height=70, key="npn_obs")

    st.markdown("---")
    if st.button("💾 Criar Projeto", type="primary", use_container_width=True, key="npn_criar"):
        if not nome:
            st.error("Preencha o Nome do Projeto.")
        elif previsto <= 0:
            st.error("Preencha o Valor Previsto.")
        else:
            pid = criar_projeto(unidade_sel, {
                "nome": nome, "tipo": tipo, "va_ggf": va,
                "responsavel": resp, "descricao": desc, "obs": obs,
                "inicio": str(data_inicio_auto) if data_inicio_auto else "",
                "termino": str(data_fim_auto) if data_fim_auto else "",
                "mes_primeiro_retorno": str(mpr),
                "previsto_unidade": previsto, "status": status,
                "check_a3": ck_a3, "check_memoria": ck_mem,
                "check_formalizado": ck_for, "ganho_unico": int(ganho_unico),
                "origem": "novo", "numero_projeto": numero_p, "lider_projeto": lider_p,
                "integrantes": integrantes_p, "revisao": revisao_p,
                "atividade_atual": "", "onde_parado": "", "data_conclusao_ativ": "",
            }, user["id"])
            if link1_tit and link1_url:
                add_link(pid, link1_tit, link1_url)
            for campo, _ in CAMPOS_A3:
                texto = st.session_state.get(f"npn_a3_{campo}", "")
                if texto:
                    salvar_a3(pid, {campo: texto})
                for img in st.session_state.get(f"npn_imgs_{campo}", []):
                    add_a3_midia(pid, campo, img["nome"], img["tipo"], img["b64"])
            for e in st.session_state.get("npn_evid", []):
                add_evidencia(pid, e["nome"], e["tipo"], e["b64"])
            for row in linhas_ativ:
                add_atividade(pid, {
                    "nome": str(row["Atividade"]).strip(),
                    "responsavel": str(row.get("Responsável") or ""),
                    "acao": str(row.get("Ação") or ""),
                    "inicio_previsto": str(_parse_date(row.get("Início Previsto")) or "") or None,
                    "termino_previsto": str(_parse_date(row.get("Término Previsto")) or "") or None,
                    "progresso_real": float(row.get("% Progresso Real") or 0),
                })
            _limpar_estado_novo_projeto()
            st.success(f"✅ Projeto **{nome}** criado! ID #{pid}. Pra continuar editando A3/Estrutura/Gantt "
                      f"depois, use o ✏️ em Minha Unidade.")
            st.rerun()

# =====================================================================
# PROJETOS APLICADOS — formulário original, disponível só até 01/01/2027
# =====================================================================
def _render_projetos_aplicados(user, colors):
    NAVY=colors.get("NAVY","#0B0F2B"); AMBER=colors.get("AMBER","#E8A838")
    GREEN=colors.get("GREEN","#1AA260"); SILVER=colors.get("SILVER","#8A9BB0")

    unidades = listar_unidades()
    nomes_u  = [u["nome"] for u in unidades]

    if user["perfil"] == "admin":
        unidade_sel = st.selectbox("Unidade:", nomes_u, key="np_uni")
    elif user["perfil"] in ("facilitador","gestor","cost_control"):
        unidade_sel = st.selectbox("Unidade:", nomes_u, key="np_uni")
        if user.get("unidade") and unidade_sel != user.get("unidade"):
            st.info(f"👁️ Você cria projetos apenas em **{user.get('unidade')}**")
    else:
        unidade_sel = user.get("unidade","")
        if unidade_sel not in nomes_u:
            st.warning("Unidade não configurada."); return

    pode = pode_editar(user, unidade_sel)
    if not pode:
        st.warning(f"⛔ Você só pode criar projetos em **{user.get('unidade','')}**.")
        return

    st.markdown("#### Identificação")
    c1,c2,c3 = st.columns(3)
    with c1: tipo = st.selectbox("Tipo *", TIPOS_PROJETO, key="np_tipo")
    with c2: va   = st.selectbox("VA / GGF / Material Auxiliar *", VA_GGF_OPTS, key="np_va")
    with c3: resp = st.text_input("Responsável", key="np_resp")

    is_extra = tipo in EXTRA_DRE_TIPOS
    if is_extra:
        st.markdown(
            f'<div style="background:#F3E8FF;border-left:3px solid #9B59B6;'
            f'border-radius:0 6px 6px 0;padding:10px 14px;font-size:11px;color:#6C3483;">'
            f'<b>↷ Extra DRE</b> — <b>{tipo}</b> gera valor operacional mas <b>não impacta o DRE</b>. '
            f'Custos valida o valor, mas não há lançamento de real mensal para este projeto. '
            f'O valor acumula mês a mês no indicador <b>Extra DRE</b> do dashboard.<br>'
            f'<span style="color:#555;">São Extra DRE: Kaizen - Custo Evitado, '
            f'Kaizen - Capital de Giro e Meta Executiva.</span></div>',
            unsafe_allow_html=True)
    else:
        st.markdown(
            f'<div style="background:#E6F4EC;border-left:3px solid {GREEN};'
            f'border-radius:0 6px 6px 0;padding:10px 14px;font-size:11px;color:#1A5C2E;">'
            f'<b>✓ Dentro do DRE</b> — <b>{tipo}</b> impacta diretamente o DRE. '
            f'Após aprovação de Custos, o valor é distribuído em 12 meses a partir do '
            f'"Ganho a partir de" e passa a ter acompanhamento de real mensal.<br>'
            f'<span style="color:#555;">São DRE: BSW, Kaizen, Kaizen - Ganho Recorrente, '
            f'Redução de Custo, Você Resolve e Estratégia Comercial.</span></div>',
            unsafe_allow_html=True)

    with st.form("form_novo", clear_on_submit=True):

        nome = st.text_input("Nome do Projeto *")
        desc = st.text_area("Descrição / Objetivo", height=70)

        st.markdown("#### Datas e Valores")
        ganho_unico = st.checkbox(
            "🎯 Ganho Único — retorno pontual, só no mês do 1º retorno",
            help="Use pra projetos raros em que o ganho inteiro acontece de uma vez, "
                 "no próprio mês de retorno — sem rateio em 12 meses. Isso evita ficar "
                 "esperando lançamento de real nos meses seguintes, que nesse caso não existem.")
        c1,c2,c3 = st.columns(3)
        with c1: inicio  = st.date_input("Data de Início do Projeto", format="DD/MM/YYYY")
        with c2: termino = st.date_input("Data de Fim do Projeto", format="DD/MM/YYYY")
        with c3: mpr     = st.date_input(
            "Ganho a partir de... *", format="DD/MM/YYYY",
            help="Mês em que o projeto começa a gerar ganho financeiro. "
                 + ("Como é Ganho Único, todo o valor entra nesse mês só."
                    if ganho_unico else
                    "A partir daqui contam 12 meses de vida útil. "
                    "Obrigatório também para projetos Extra DRE."))

        previsto = st.number_input(
            "Valor Previsto (R$) *",
            min_value=0.0, step=1000.0, format="%.2f",
            help="Valor estimado pela unidade. "
                 + ("Ganho Único: valor inteiro lançado no mês de retorno, sem rateio."
                    if ganho_unico else
                    "Distribuído em 12 meses a partir do 1º retorno. "
                    + ("Para Extra DRE, acumula no indicador Extra DRE mês a mês." if is_extra else
                       "Após validação de Custos, o valor calculado substituirá este nas projeções.")))

        st.markdown("#### Links e Evidências (SharePoint / OneDrive)")
        st.caption("Cole os links completos (https://...). Você pode adicionar mais depois.")
        c1,c2 = st.columns([2,4])
        with c1: link1_tit = st.text_input("Nome 1", placeholder="ex: A3 do Projeto")
        with c2: link1_url = st.text_input("URL 1", placeholder="https://...")
        c1,c2 = st.columns([2,4])
        with c1: link2_tit = st.text_input("Nome 2", placeholder="ex: Memória de Cálculo")
        with c2: link2_url = st.text_input("URL 2", placeholder="https://...")
        c1,c2 = st.columns([2,4])
        with c1: link3_tit = st.text_input("Nome 3", placeholder="ex: Planilha de Apoio")
        with c2: link3_url = st.text_input("URL 3", placeholder="https://...")

        st.markdown("#### Acompanhamento")
        status = st.selectbox("Status", STATUS_OPTS)
        st.caption("Os campos abaixo são opcionais — preenchidos nos acompanhamentos semanais.")
        c1,c2,c3 = st.columns(3)
        with c1: ativ     = st.text_input("Atual Atribuição", placeholder="Atividade em andamento...")
        with c2: resp_ativ = st.text_input("Responsável da Atribuição", placeholder="Nome...")
        with c3: dt_ativ   = st.text_input("Data Final da Atribuição", placeholder="ex: 08/2026")

        st.markdown("#### Checklist")
        st.caption("Quando os 3 estiverem marcados, o projeto vai para validação de Custos.")
        c1,c2,c3 = st.columns(3)
        with c1: ck_a3  = st.checkbox("A3 e Plano de Projeto desenvolvido")
        with c2: ck_mem = st.checkbox("Memória de Cálculo desenvolvida")
        with c3: ck_for = st.checkbox("Formalizado com Custos")
        if ck_a3 and ck_mem and ck_for:
            st.success("✅ Checklist completo — será encaminhado para validação de Custos.")

        st.markdown("#### Observações")
        obs = st.text_area("", height=80, label_visibility="collapsed",
                            placeholder="Observações gerais sobre o projeto...")

        st.markdown("---")
        salvar = st.form_submit_button("💾 Cadastrar Projeto",
                                       use_container_width=True, type="primary")

    if salvar:
        if not nome:
            st.error("Preencha o Nome do Projeto.")
        elif previsto <= 0:
            st.error("Preencha o Valor Previsto.")
        else:
            pid = criar_projeto(unidade_sel, {
                "nome": nome, "tipo": tipo, "va_ggf": va,
                "responsavel": resp, "descricao": desc, "obs": obs,
                "inicio": str(inicio), "termino": str(termino),
                "mes_primeiro_retorno": str(mpr),
                "previsto_unidade": previsto, "status": status,
                "atividade_atual": ativ,
                "data_conclusao_ativ": dt_ativ,
                "onde_parado": resp_ativ,
                "check_a3": ck_a3, "check_memoria": ck_mem,
                "check_formalizado": ck_for, "ganho_unico": int(ganho_unico),
                "origem": "aplicado",
            }, user["id"])
            for tit, url in [(link1_tit,link1_url),(link2_tit,link2_url),(link3_tit,link3_url)]:
                if tit and url:
                    add_link(pid, tit, url)
            st.success(f"✅ Projeto **{nome}** cadastrado! ID #{pid}")
            st.rerun()

# =====================================================================
def render(user, **colors):
    st.markdown('<span class="st">Projetos</span>', unsafe_allow_html=True)

    aplicados_disponivel = date.today() < DATA_FIM_PROJETOS_APLICADOS
    if aplicados_disponivel:
        tab_novo, tab_aplicados = st.tabs(["🆕 Novo Projeto", "📝 Projetos Aplicados"])
        with tab_novo:
            _render_novo_projeto(user, colors)
        with tab_aplicados:
            _render_projetos_aplicados(user, colors)
    else:
        st.caption("O formulário 'Projetos Aplicados' (usado pra cadastrar o histórico) não está mais "
                  "disponível a partir de 01/01/2027 — todo projeto novo nasce no formato completo, "
                  "com A3, Estrutura e Gantt.")
        _render_novo_projeto(user, colors)
